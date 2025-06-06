import pandas as pd
from openpyxl import load_workbook # Not strictly needed for writing new, but good to know
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

import re # For case-insensitive "test" matching
import os
import json
import numpy as np

with open("AOP Automation Scripts/input_data/keyword_groups.json", "r", encoding="utf-8") as f:
    loaded = json.load(f)
    red_keywords_group = loaded["red"]
    red_keywords_complex = loaded["red_complex"]
    green_keywords_group = loaded["green"]
    green_keywords_complex = loaded["green_complex"]


#red_keywords_group = ["MEMORY", "SIP", "FPS", "Molded MEMS", "3O "] # Case-sensitive as per examples
#green_keywords_group = ["CABGA", "BGA", "SCSP"]              # Case-sensitive
#green_keywords_complex = ["CA "]

def highlight_rows(column_name : str, df : pd.DataFrame):
    
    # Define keywords for red highlighting

    # Define fills
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # Light Yellow
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green


    # --- Logic to determine cell colors ---
    # We'll store color instructions: (row_index, col_index, fill_object)
    cell_formats_to_apply = []
    remove_activated = False
    plant_col_idx = df.columns.get_loc(column_name) + 1 # 1-based index for openpyxl

    for r_idx, row in df.iterrows():
        # openpyxl rows are 1-based, and ExcelWriter usually writes headers, so add 2
        # (1 for 1-based, 1 for header row)
        excel_row_num = r_idx + 2

        if excel_row_num == 695:
            pass

        current_row_override = False
        cell_value_plant = str(row[column_name]) # Ensure it's a string

        # --- Rule 1 & 2: Red Highlighting for whole row based on 'Plant' column ---
        is_red = any(re.search(keyword, cell_value_plant, re.IGNORECASE) for keyword in red_keywords_group)
        is_green = any(re.search(keyword, cell_value_plant, re.IGNORECASE) for keyword in green_keywords_group)
        
        #RED Rule
        if is_red:
            remove_activated = True
            current_row_override = True
            for c_idx in range(1, len(df.columns) + 1): # Apply to all cells in the row
                cell_formats_to_apply.append({'row': excel_row_num, 'column': c_idx, 'fill': red_fill, 'overwrite_previous_for_cell': True})

        #Green Rule
        elif is_green:
            remove_activated = False
            current_row_override = True
            for c_idx in range(1, len(df.columns) + 1):
                cell_formats_to_apply.append({'row': excel_row_num, 'column': c_idx, 'fill': green_fill})

        if re.search(r'test', cell_value_plant, re.IGNORECASE):
            if remove_activated and not current_row_override:
                for c_idx in range(1, len(df.columns) + 1):
                    cell_formats_to_apply.append({'row': excel_row_num, 'column': c_idx, 'fill': yellow_fill})
            elif not remove_activated:
                for c_idx in range(1, len(df.columns) + 1):
                    cell_formats_to_apply.append({'row': excel_row_num, 'column': c_idx, 'fill': green_fill})
            # Override for specific test use cases 
            if any(re.search(keyword, str(row['Unnamed: 3']), re.IGNORECASE) for keyword in green_keywords_complex):
                for c_idx in range(1, len(df.columns) + 1): # Apply to all cells in the row
                    cell_formats_to_apply.append({'row': excel_row_num, 'column': c_idx, 'fill': green_fill})
            if any(re.search(keyword, str(row['Unnamed: 3']), re.IGNORECASE) for keyword in red_keywords_complex):
                for c_idx in range(1, len(df.columns) + 1): # Apply to all cells in the row
                    cell_formats_to_apply.append({'row': excel_row_num, 'column': c_idx, 'fill': red_fill, 'overwrite_previous_for_cell': True})

    return cell_formats_to_apply, df

def remove_rows(column_name : str, df : pd.DataFrame):

    # Define fills
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # Light Yellow
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green


    # --- Logic to determine cell colors ---
    # We'll store color instructions: (row_index, col_index, fill_object)
    remove_activated = False
    remove_index = []

    for r_idx, row in df.iterrows():
        # openpyxl rows are 1-based, and ExcelWriter usually writes headers, so add 2
        # (1 for 1-based, 1 for header row)
        excel_row_num = r_idx + 2
        current_row_override = False
        cell_value_plant = str(row[column_name]) # Ensure it's a string

        # --- Rule 1 & 2: Red Highlighting for whole row based on 'Plant' column ---
        is_red = any(re.search(keyword, cell_value_plant, re.IGNORECASE) for keyword in red_keywords_group)
        is_green = any(re.search(keyword, cell_value_plant, re.IGNORECASE) for keyword in green_keywords_group)
            
        if is_red:
            remove_activated = True
            current_row_override = True
            remove_index.append(r_idx)

        elif is_green:
            remove_activated = False
            current_row_override = True
        else:
            if re.search(r'test', cell_value_plant, re.IGNORECASE):
                if remove_activated and not current_row_override:
                    remove_index.append(r_idx)
                elif not remove_activated:
                    pass
                elif current_row_override:
                    print(f"ERROR! Override in the Test on row {excel_row_num}")
            if any(re.search(keyword, str(row['Unnamed: 3']), re.IGNORECASE) for keyword in green_keywords_complex):
                if r_idx in remove_index: 
                    remove_index.remove(r_idx)
            if any(re.search(keyword, str(row['Unnamed: 3']), re.IGNORECASE) for keyword in red_keywords_complex):
                if r_idx in remove_index: 
                    remove_index.remove(r_idx)

    #Flip the remove index and then remove unecessary rows from Dataframe                 
    remove_index.reverse()
    df = df.drop(index=remove_index)#.reset_index(drop=True)

    return remove_index, df 

def modify_headers(df : pd.DataFrame) -> pd.DataFrame:
    df.columns = [''.join(col.split(' ')[0]) for col in df.columns]
    column_names = []
    for i in range(df.shape[1]):
        column_names.append(df.columns[i] + " " + df.iloc[0, i])
    df.columns = column_names
    
    return df.drop(index=0)
    
    
def pivot_table(df : pd.DataFrame, output_filename : str):
    df.iloc[:, 4:] = df.iloc[:, 4:].apply(pd.to_numeric, errors='coerce').fillna(0)
    
    aggregate_vals = [col for col in df.columns[4:] if pd.api.types.is_numeric_dtype(df[col])]

    pivot_df = pd.pivot_table(
        df,
        values=aggregate_vals,      # What to aggregate
        index=' Legal Name',      # Rows of the pivot table
        aggfunc='sum',       # How to aggregate (sum, mean, count, etc.)
        fill_value=0         # Value to fill for missing combinations
    )
    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_df.to_excel(writer, sheet_name='Pivot Table')

            # --- 5. Apply Styling (Optional) ---
            workbook = writer.book # Get the openpyxl workbook object
            worksheet_pivot = workbook['Pivot Table'] # Get the specific sheet

            # --- Define Styles ---
            header_font = Font(bold=True, color="000000") # Black font
            header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid") # Light green
            thin_border_side = Side(border_style="thin", color="000000")
            header_border = Border(top=thin_border_side, left=thin_border_side, right=thin_border_side, bottom=thin_border_side)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

            # --- Apply Header Styling ---
            # Pandas writes the index names and column names in the first row.
            # Index names are in columns 1 to n_levels
            # Column names start after index columns
            n_levels = pivot_df.index.nlevels
            for col_idx in range(1, worksheet_pivot.max_column + 1):
                cell = worksheet_pivot.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = header_alignment
            
            # Column Dimension Resizing 
            worksheet_pivot.column_dimensions[get_column_letter(1)].width = 18 # Region
            worksheet_pivot.column_dimensions[get_column_letter(2)].width = 18 # Sales Rep

            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light red
            dark_red_font = Font(color="9C0006") # Dark red font

            # Rule: cell value > 400
            conditional_rule = CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, font=dark_red_font, fill=red_fill)
            worksheet_pivot.conditional_formatting.add('$D3:$Z100', conditional_rule)


       

    except ImportError:
        print("openpyxl is not installed. Styling cannot be applied this way.")
        print("Please install it: pip install openpyxl pandas")
        # Fallback to basic Excel writing without advanced styling (Pandas default if openpyxl not specified as engine)
        with pd.ExcelWriter(output_filename) as writer: # engine defaults to openpyxl if available
            df.to_excel(writer, sheet_name='Raw_Data', index=False)
            pivot_df.to_excel(writer, sheet_name='Pivot Table', index=True)
        print(f"\nExcel file '{output_filename}' created successfully (basic).")
    except Exception as e:
        print(f"An error occurred: {e}")




def apply_conditional_formatting(input_excel_path, output_excel_path, task='remove', column_name="Unnamed: 2", sheet_name="Sheet1"):
    try:
        df = pd.read_excel(input_excel_path, sheet_name=sheet_name, keep_default_na=False)
    except FileNotFoundError:
        print(f"Error: Input file '{input_excel_path}' not found.")
        return
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    if column_name not in df.columns:
        print(f"Error: Column '{column_name}' not found in the sheet.")
        print(f"Available columns are: {df.columns.tolist()}")
        return

    # Determine if file exists
    file_exists = os.path.exists(output_excel_path)

    if task == 'both':
        print("Running Both Commands")
        # Get both DataFrames and formatting instructions
        remove_index, remove_df = remove_rows(column_name, df)
        remove_df.iloc[0,0:4]=['Date', 'Legal Name', 'Pkg', 'PDL']
        remove_df.rename(columns={'Unnamed: 0' : ' ', 'Unnamed: 1' : ' ', 'Unnamed: 2' : ' ', 'Unnamed: 3' : ' '}, inplace=True)
        cell_formats_to_apply, highlight_df = highlight_rows(column_name, df)
        try:
            if file_exists:
                writer = pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
            else:
                writer = pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='w')
            with writer:
                highlight_df.to_excel(writer, sheet_name='highlight', index=False)
                remove_df.to_excel(writer, sheet_name='remove', index=False)
            wb = load_workbook(output_excel_path)
            ws = wb['highlight']
            applied_fills = {}
            for fmt_instruction in cell_formats_to_apply:
                cell_coord = (fmt_instruction['row'], fmt_instruction['column'])
                if fmt_instruction.get('overwrite_previous_for_cell') or cell_coord not in applied_fills:
                    cell = ws.cell(row=fmt_instruction['row'], column=fmt_instruction['column'])
                    cell.fill = fmt_instruction['fill']
                    applied_fills[cell_coord] = fmt_instruction['fill']
            wb.save(output_excel_path)
            print(f"Successfully wrote both sheets and applied formatting to '{output_excel_path}'")
        except Exception as e:
            print(f"Error writing Excel file or applying styles: {e}")
            import traceback
            traceback.print_exc()
        return

    if task == 'remove':
        print("running remove command")
        remove_index, remove_df = remove_rows(column_name, df)
        remove_df.iloc[0,0:4]=['Date', 'Legal Name', 'Pkg', 'PDL']
        remove_df.rename(columns={'Unnamed: 0' : ' ', 'Unnamed: 1' : ' ', 'Unnamed: 2' : ' ', 'Unnamed: 3' : ' '}, inplace=True)
        remove_df = modify_headers(remove_df)

        try:
            if file_exists:
                writer = pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
            else:
                writer = pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='w')
            with writer:
                remove_df.to_excel(writer, sheet_name='remove', index=False)
            print(f"Successfully updated 'remove' sheet in '{output_excel_path}'")
        except Exception as e:
            print(f"Error writing Excel file: {e}")
            import traceback
            traceback.print_exc()
        
        try: 
            pivot_table(remove_df, output_excel_path)
        except Exception as e:
            print(f"Error writing Pivot Table: {e}")
            import traceback
            traceback.print_exc()

    if task == 'highlight':
        print("running Highlight Command")
        cell_formats_to_apply, highlight_df = highlight_rows(column_name, df)
        try:
            if file_exists:
                writer = pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
            else:
                writer = pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='w')
            with writer:
                highlight_df.to_excel(writer, sheet_name='highlight', index=False)
            wb = load_workbook(output_excel_path)
            ws = wb['highlight']
            applied_fills = {}
            for fmt_instruction in cell_formats_to_apply:
                cell_coord = (fmt_instruction['row'], fmt_instruction['column'])
                if fmt_instruction.get('overwrite_previous_for_cell') or cell_coord not in applied_fills:
                    cell = ws.cell(row=fmt_instruction['row'], column=fmt_instruction['column'])
                    cell.fill = fmt_instruction['fill']
                    applied_fills[cell_coord] = fmt_instruction['fill']
            wb.save(output_excel_path)
            print(f"Successfully updated and formatted 'highlight' sheet in '{output_excel_path}'")
        except Exception as e:
            print(f"Error writing Excel file or applying styles: {e}")
            import traceback
            traceback.print_exc()



# --- Example Usage ---
if __name__ == "__main__":
    input_file = "AOP Automation Scripts/output_data/converted_report.xlsx"
    output_file = "AOP Automation Scripts/output_data/formatted_report.xlsx"
    apply_conditional_formatting(input_file, output_file, task = "remove")